import io
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render

import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from achats.models import Achat
from ventes.models import Vente
from partenaires.models import Client, Fournisseur
from stock.services import calculer_stock

BLUE = colors.HexColor("#2955e0")


@login_required
def index(request):
    return render(request, "exports/index.html")


def _dataset(nom):
    if nom == "achats":
        headers = ["N° Bon", "Facture", "Fournisseur", "Matière", "Quantité (Kg)", "Prix unitaire", "Montant", "Date", "Statut"]
        rows = [[a.numero_bon, a.numero_facture, str(a.fournisseur), str(a.materiau), float(a.quantite),
                  float(a.prix_unitaire), float(a.montant_total), str(a.date_achat), a.statut_paiement]
                 for a in Achat.objects.select_related("fournisseur", "materiau").all()]
        return "Achats", headers, rows

    if nom == "ventes":
        headers = ["Facture", "N° Bon", "Client", "Matière", "Quantité (Kg)", "Prix unitaire", "Montant", "Date", "Statut"]
        rows = [[v.numero_facture, v.numero_bon, str(v.client), str(v.materiau), float(v.quantite),
                  float(v.prix_unitaire), float(v.montant_total), str(v.date_vente), v.statut_paiement]
                 for v in Vente.objects.select_related("client", "materiau").all()]
        return "Ventes", headers, rows

    if nom == "clients":
        headers = ["Nom", "Téléphone", "Ville", "Nb ventes", "Total acheté", "Total payé", "Restant"]
        rows = [[c.nom, c.telephone, c.ville, c.nombre_ventes, float(c.total_achete), float(c.total_paye), float(c.montant_restant)]
                 for c in Client.objects.all()]
        return "Clients", headers, rows

    if nom == "fournisseurs":
        headers = ["Nom", "Téléphone", "Ville", "Nb achats", "Total acheté", "Total payé", "Reste à payer"]
        rows = [[f.nom, f.telephone, f.ville, f.nombre_achats, float(f.total_achete), float(f.total_paye), float(f.reste_a_payer)]
                 for f in Fournisseur.objects.all()]
        return "Fournisseurs", headers, rows

    if nom == "stock":
        headers = ["Matière", "Quantité achetée (Kg)", "Quantité vendue (Kg)", "Stock restant (Kg)"]
        rows = [[s["materiau"].nom, float(s["quantite_achetee_kg"]), float(s["quantite_vendue_kg"]), float(s["stock_kg"])]
                 for s in calculer_stock()]
        return "Stock", headers, rows

    return None, [], []


@login_required
def export_excel(request, nom):
    titre, headers, rows = _dataset(nom)
    if titre is None:
        return HttpResponse("Jeu de données inconnu", status=404)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titre[:31]

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2955E0", end_color="2955E0", fill_type="solid")

    for row in rows:
        ws.append(row)

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{titre}.xlsx"'
    return response


@login_required
def export_pdf(request, nom):
    titre, headers, rows = _dataset(nom)
    if titre is None:
        return HttpResponse("Jeu de données inconnu", status=404)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm, leftMargin=12*mm, rightMargin=12*mm)
    styles = getSampleStyleSheet()

    elements = [Paragraph(f"Export - {titre}", styles["Title"]), Spacer(1, 6*mm)]

    table_data = [headers] + [[str(v) for v in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BLUE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e3e8f1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6fb")]),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{titre}.pdf"'
    return response
